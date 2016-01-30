import openpyxl as xl
from openpyxl.cell import get_column_letter


def workbook(path=None):

    if path:
        try:
            return xl.load_workbook(path)
        except:
            wb = xl.Workbook()
            append_row(wb, ['Print No.', 'File Name', 'Reason', 'Requested By',
                            'Recipicant', 'Watermarked by', 'Printed By', 'Paper size', 'Date Sent', 'Location On Server'])
            return wb
    else:
        wb = xl.Workbook()
        append_row(wb, ['Print No.', 'File Name', 'Reason', 'Requested By',
                        'Recipicant', 'Watermarked by', 'Printed By', 'Paper size', 'Date Sent', 'Location On Server'])
        return wb


def rows_columns(wb):

    sheet = wb.active

    return len(sheet.rows), len(sheet.columns)

def append_row(wb, data):

    sheet = wb.active
    r, c = rows_columns(wb)
    r += 1
    for c, el in enumerate(data):

        sheet[get_column_letter(c+1)+str(r)] = el


if __name__ == '__main__':

    wb = workbook(r'C:\Users\Pysenberg\Desktop\bok.xlsx')
    append_row(wb, ['a','b','c','d'])
    wb.save(r'C:\Users\Pysenberg\Desktop\bok.xlsx')
